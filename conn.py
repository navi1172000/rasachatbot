import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('event.db')
wb = load_workbook('events.xlsx')
ws = wb['items']

conn.execute("create table if not exists events (cityname text, event text, price int)")

for i in range(1,6):
    temp_str = "insert into events (cityname, event, price) values ('{0}', '{1}', '{2}')".format(ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value)
    conn.execute(temp_str)

conn.commit()

content = conn.execute("select * from events")
for i in content:
    print(i)
